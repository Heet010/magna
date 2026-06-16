"""Render the executive summary as Markdown, Excel and a verification table."""
from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import config
from src.schema import CompanyExtraction, PeriodMetrics

SUBSTITUTED_MARK = " †"  # appended to values drawn from an equivalent KPI


def _metric(period: PeriodMetrics, key: str):
    for m in period.metrics:
        if m.metric_key == key:
            return m
    return None


def _cell_text(period: PeriodMetrics, key: str) -> str:
    m = _metric(period, key)
    if m is None:
        return "—"
    val = m.value
    if m.status == "substituted":
        val += SUBSTITUTED_MARK
    return val


def _periods(extractions: list[CompanyExtraction]):
    """Flatten to ordered (company, period_key, PeriodMetrics) tuples -> 6 columns."""
    cols = []
    for ex in extractions:
        cols.append((ex.company, "FY2025", ex.full_year))
        cols.append((ex.company, "Q1-2026", ex.last_quarter))
    return cols


def render_markdown(extractions: list[CompanyExtraction]) -> str:
    cols = _periods(extractions)
    short = {
        "BMW Group": "BMW",
        "Mercedes-Benz Group": "Mercedes-Benz",
        "Volkswagen Group": "Volkswagen",
    }

    header = ["Metric"] + [
        f"{short.get(c, c)} — {'FY 2025' if p == 'FY2025' else 'Q1 2026'}"
        for c, p, _ in cols
    ]
    lines = [
        "# Executive Summary — Automotive OEM Financials",
        "",
        f"*Generated {date.today().isoformat()} by an AI agent (OpenAI {config.AGENT_MODEL}) "
        "from the source reports only.*",
        "",
        "| " + " | ".join(header) + " |",
        "| " + " | ".join(["---"] * len(header)) + " |",
    ]
    for m in config.METRICS:
        row = [m["label"]] + [
            _cell_text(pm, m["key"]).replace("|", "\\|") for _, _, pm in cols
        ]
        lines.append("| " + " | ".join(row) + " |")

    lines += ["", f"† = company's equivalent/substituted KPI (not an exact term match). "
              "“Not Reported” = absent from the quarterly report; “N/A” = not disclosed.", ""]

    lines += ["## Substituted & equivalent KPIs", ""]
    for ex in extractions:
        lines.append(f"**{ex.company}** — {ex.summary_note.strip()}")
        subs = []
        for period_key, pm in (("FY 2025", ex.full_year), ("Q1 2026", ex.last_quarter)):
            for met in pm.metrics:
                if met.status == "substituted" and met.note.strip():
                    subs.append(
                        f"  - {config.METRIC_LABELS.get(met.metric_key, met.metric_key)} "
                        f"({period_key}): used *{met.source_term}* — {met.note.strip()}"
                    )
        lines += subs + [""]

    return "\n".join(lines)


def render_verification(extractions: list[CompanyExtraction]) -> str:
    """A flat audit table: every figure with the KPI term, status and source page."""
    lines = [
        "# Verification — figure provenance",
        "",
        "Each extracted figure with the exact KPI term the company used, its status, "
        "and the **PDF page** the agent read it from. Open the matching report and jump "
        "to that page to confirm the number.",
        "",
        "| Company | Period | Metric | Value | KPI term used | Status | Page |",
        "| --- | --- | --- | --- | --- | --- | --- |",
    ]
    for ex in extractions:
        for label, pm in (("FY 2025", ex.full_year), ("Q1 2026", ex.last_quarter)):
            for m in config.METRICS:
                met = _metric(pm, m["key"])
                if met is None:
                    continue
                row = [
                    ex.company, label, m["label"],
                    met.value, met.source_term or "—", met.status,
                    met.source_page or "—",
                ]
                lines.append("| " + " | ".join(c.replace("|", "\\|") for c in row) + " |")
    lines.append("")
    return "\n".join(lines)


def print_verification(extractions: list[CompanyExtraction]) -> None:
    """Compact console spot-check: value [page] per metric, per company/period."""
    for ex in extractions:
        print(f"\n=== {ex.company} ===")
        for label, pm in (("FY2025", ex.full_year), ("Q1-2026", ex.last_quarter)):
            print(f"  [{label}]")
            for m in config.METRICS:
                met = _metric(pm, m["key"])
                if met is None:
                    continue
                page = f"p.{met.source_page}" if met.source_page else "—"
                flag = "*" if met.status == "substituted" else " "
                print(f"   {flag} {m['label']:<38} {met.value:<22} [{page}]")


_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_GROUP_FILL = PatternFill("solid", fgColor="374151")
_SUB_FILL = PatternFill("solid", fgColor="FEF3C7")     # substituted
_NA_FILL = PatternFill("solid", fgColor="F3F4F6")      # N/A / not reported
_THIN = Side(style="thin", color="D1D5DB")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def render_excel(extractions: list[CompanyExtraction], path: Path) -> None:
    cols = _periods(extractions)
    wb = Workbook()
    ws = wb.active
    ws.title = "Executive Summary"

    # Row 1: company group header (merged over its two period columns)
    ws.cell(1, 1, "Metric")
    ws.cell(2, 1, "")
    for idx, ex in enumerate(extractions):
        c0 = 2 + idx * 2
        ws.cell(1, c0, ex.company)
        ws.merge_cells(start_row=1, start_column=c0, end_row=1, end_column=c0 + 1)
        ws.cell(2, c0, "FY 2025")
        ws.cell(2, c0 + 1, "Q1 2026")

    # Data rows
    for r, m in enumerate(config.METRICS, start=3):
        ws.cell(r, 1, m["label"])
        for ci, (_, _, pm) in enumerate(cols, start=2):
            met = _metric(pm, m["key"])
            cell = ws.cell(r, ci, met.value if met else "—")
            if met and met.status == "substituted":
                cell.fill = _SUB_FILL
            elif met and met.status in ("not_available", "not_reported"):
                cell.fill = _NA_FILL

    # Styling
    for c in range(1, 2 + len(cols)):
        ws.cell(1, c).font = Font(bold=True, color="FFFFFF")
        ws.cell(1, c).fill = _HEADER_FILL
        ws.cell(2, c).font = Font(bold=True, color="FFFFFF")
        ws.cell(2, c).fill = _GROUP_FILL
    for row in ws.iter_rows(min_row=1, max_row=2 + len(config.METRICS), max_col=1 + len(cols)):
        for cell in row:
            cell.border = _BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 34
    for c in range(2, 2 + len(cols)):
        ws.column_dimensions[get_column_letter(c)].width = 22
    ws.freeze_panes = "B3"

    # Notes sheet
    notes = wb.create_sheet("Notes & Substitutions")
    notes.cell(1, 1, "Company").font = Font(bold=True)
    notes.cell(1, 2, "Summary").font = Font(bold=True)
    nr = 2
    for ex in extractions:
        notes.cell(nr, 1, ex.company)
        notes.cell(nr, 2, ex.summary_note.strip())
        nr += 1
    nr += 1
    notes.cell(nr, 1, "Metric").font = Font(bold=True)
    notes.cell(nr, 2, "Detail").font = Font(bold=True)
    nr += 1
    for ex in extractions:
        for period_key, pm in (("FY 2025", ex.full_year), ("Q1 2026", ex.last_quarter)):
            for met in pm.metrics:
                if met.status in ("substituted", "not_reported", "not_available") and (
                    met.note.strip() or met.status != "reported"
                ):
                    label = config.METRIC_LABELS.get(met.metric_key, met.metric_key)
                    detail = f"[{met.status}] {met.source_term or ''} {met.note}".strip()
                    notes.cell(nr, 1, f"{ex.company} — {label} ({period_key})")
                    notes.cell(nr, 2, detail)
                    nr += 1
    notes.column_dimensions["A"].width = 48
    notes.column_dimensions["B"].width = 90

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
